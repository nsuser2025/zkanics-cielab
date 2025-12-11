import streamlit as st
import string
import pandas as pd
from openpyxl import load_workbook
import numpy as np

def sanitize_for_csv_injection(df):
    for col in df.columns:
        if df[col].dtype == 'object':
           df[col] = (
                df[col]
                .fillna("")
                .astype(str)
                .str.replace(r'^([=+\-@])', r"'\1", regex=True)
            )
    return df

def mkcsv_gui(uploaded_file):

    try:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active
    except Exception as e:
        st.error(f"Excel/CSVファイルの読み込み中にエラーが発生しました: {e}")
        st.stop()
      
    st.markdown("---")  
    df_orig = pd.read_excel(uploaded_file, header=None)
    df_orig.columns = list(string.ascii_uppercase[:len(df_orig.columns)])
    df_orig.index = range(1, len(df_orig) + 1)
    df_orig_safe = sanitize_for_csv_injection(df_orig.copy()) 
    # 追加：列名をすべてstrに統一 2025/12/03 START
    df_orig_safe.columns = df_orig_safe.columns.map(str)
    # 2025/12/03 END
    st.dataframe(df_orig_safe) 
    st.markdown("---")
    
    st.markdown("**各データのExcel範囲をA1:A10のように入力してください。（1列のみ対応）**")
    file_input = st.text_input("ファイル名範囲", value="A1:A10", key="k_file")
    exam_input = st.text_input("試験範囲", value="B1:B10", key="k_exam")
    face_input = st.text_input("測定面範囲", value="C1:C10", key="k_face")
    cath_input = st.text_input("正極範囲", value="D1:D10", key="k_cath")
    mesu_input = st.text_input("測定範囲", value="E1:E10", key="k_mesu")
    elec_input = st.text_input("電解液範囲", value="F1:F10", key="k_elec")
    magn_input = st.text_input("倍率範囲", value="G1:G10", key="k_magn")
    
    range_inputs = {
        "ファイル名": file_input,
        "試験": exam_input,
        "測定面": face_input,
        "正極": cath_input,
        "測定": mesu_input,
        "電解液": elec_input,
        "倍率": magn_input,
    }

    
    # 範囲ごとの値を抽出する関数 (1列範囲に限定し、エラー時はNoneを返す)
    def extract_range_data(range_str):
        if not range_str:
            return None, "範囲が空です"
        try:
            cells = ws[range_str]
            data_list = [c[0].value for c in cells]
            return data_list, None
        except Exception:
            return None, f"'{range_str}' の範囲指定が無効です。"

    if st.button("CSVファイルを生成"):
       extracted_data = {}
       error_messages = []

       for col_name, range_str in range_inputs.items():
           data_list, error = extract_range_data(range_str)
           if error:
                 error_messages.append(f"列 '{col_name}': {error}")
           extracted_data[col_name] = data_list

       if error_messages:
          st.error("以下のエラーのため、DataFrameを生成できませんでした。")
          for msg in error_messages:
              st.write(f"- {msg}")
          st.stop()
        
    
       lengths = {name: len(lst) for name, lst in extracted_data.items() if lst is not None}
       unique_lengths = set(lengths.values())

       if len(unique_lengths) > 1:
          st.error("抽出したリストの長さが一致しません")
          st.warning("DataFrameの列として結合するには、全てのリストが同じ長さである必要があります。")
          st.json(lengths)
          st.stop()
        
       if not unique_lengths:
          st.warning("全ての範囲が空でした。DataFrameを生成できません")
          st.stop()

       df_out = pd.DataFrame(extracted_data)
       st.success(f"選択範囲を結合したCSVを生成しました。（{list(unique_lengths)[0]}行）")
       df_safe = sanitize_for_csv_injection(df_out.copy()) 
       # 追加：列名をすべて str に統一 2025/12/03 START
       df_safe.columns = df_safe.columns.map(str) 
       # 2025/12/03 END
       st.dataframe(df_safe)
       return df_out

# MODULE ERROR MESSAGE
if __name__ == "__main__":
   raise RuntimeError("Do not run this file directly; use it as a module.")
